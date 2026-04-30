import logging
import re
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

from core.afip_tipos import AFIP_TIPOS
from core.comparator import conciliar
from core.dtype_detector import detect_column_type, VALID_TYPES
from core.rule_labels import RULE_LABELS, RULE_LABELS_INV
from core.rules import apply_rule
from ui.column_mapper import column_mapper
from ui.rule_builder import rule_builder
from utils.file_loader import accepted_extensions, get_excel_sheets, load_dataframe

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

st.set_page_config(page_title="Conciliador de Planillas - IT Concesionaria", layout="wide")
st.title("Conciliador General de Planillas")
st.markdown("Cargá entre 1 y 4 planillas para filtrar datos o buscar coincidencias.")

TYPE_LABELS = {"str": "Texto", "int": "Número entero", "float": "Número decimal", "date": "Fecha"}
CALC_OPS = ["×", "+", "-", "÷"]

# ── Inicializar estado ─────────────────────────────────────────────────────────
if "n_tables" not in st.session_state:
    st.session_state.n_tables = 2


# ── Helpers ────────────────────────────────────────────────────────────────────

_DIFF_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def _highlight_diff_cells(ws) -> None:
    """Aplica fondo amarillo claro a las celdas _A y _B en la hoja Diferencias."""
    headers = [cell.value for cell in ws[1]]
    diff_col_indices = {
        i + 1  # openpyxl usa índice 1-based
        for i, h in enumerate(headers)
        if h and (str(h).endswith("_A") or str(h).endswith("_B"))
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in diff_col_indices:
                cell.fill = _DIFF_FILL


def _render_file_loader(label: str, col_key: str):
    file = st.file_uploader(label, type=accepted_extensions(), key=f"file_{col_key}")
    if file is None:
        return None, "."
    decimal = "."
    if file.name.lower().endswith(".csv"):
        decimal = st.radio(
            "Separador decimal", [".", ","], horizontal=True, key=f"decimal_{col_key}",
            format_func=lambda x: f'Punto "." (ej: 1.50)' if x == "." else f'Coma "," (ej: 1,50)',
        )
    sheet = None
    if file.name.lower().endswith((".xlsx", ".xls", ".xlsm", ".xlsb")):
        sheets = get_excel_sheets(file.getvalue(), file.name)
        sheet = st.selectbox("Seleccionar hoja", sheets, key=f"sheet_{col_key}")
        if len(sheets) > 1:
            st.caption("📄 Este archivo tiene múltiples hojas")
    df = load_dataframe(file.getvalue(), file.name, sheet=sheet, decimal=decimal)
    st.caption(f"✅ {len(df):,} filas · {len(df.columns)} columnas")
    return df, decimal


def _render_type_overrides(df, detected: dict, state_key: str, file_key: str):
    if st.session_state.get(f"_fk_{state_key}") != file_key:
        st.session_state[f"_fk_{state_key}"] = file_key
        st.session_state[state_key] = dict(detected)
    overrides = st.session_state[state_key]
    for i in range(0, len(df.columns), 4):
        row_cols = list(df.columns)[i:i + 4]
        ui_cols = st.columns(4)
        for j, col_name in enumerate(row_cols):
            with ui_cols[j]:
                current = overrides.get(col_name, detected[col_name])
                new_type = st.selectbox(
                    f"**{col_name}**", VALID_TYPES, index=VALID_TYPES.index(current),
                    format_func=lambda t, c=col_name: (
                        f"{TYPE_LABELS[t]} ✏️" if t != detected.get(c) else TYPE_LABELS[t]
                    ),
                    key=f"{state_key}_{col_name}",
                    help=f"Detectado: {TYPE_LABELS[detected[col_name]]}",
                )
                overrides[col_name] = new_type
    return overrides


def _build_compound_code(df: pd.DataFrame, tipo_col: str, pv_col: str, nro_col: str,
                          output_col: str) -> pd.DataFrame:
    """Construye código de comprobante: pv.zfill(4) + nro.zfill(8) + letra AFIP.

    Extrae el código del tipo_col parseando 'N - Descripción' → N,
    normaliza a 3 dígitos y busca en AFIP_TIPOS.
    """
    result = df.copy()

    def _safe_zfill(val, width):
        try:
            return str(int(float(str(val).replace(",", ".").strip()))).zfill(width)
        except (ValueError, TypeError):
            return "0" * width

    codigos_raw = result[tipo_col].astype(str).str.extract(r"^(\d+)\s*(?:-|$)")[0].str.strip()
    # Normalizar a 3 dígitos para buscar en el diccionario (ej: "81" → "081", "1" → "001")
    codigos_norm = codigos_raw.apply(
        lambda x: str(int(x)).zfill(3) if pd.notna(x) and str(x).strip().isdigit() else ""
    )
    letras = codigos_norm.map(AFIP_TIPOS).fillna("")

    pv_padded  = result[pv_col].apply(lambda x: _safe_zfill(x, 4))
    nro_padded = result[nro_col].apply(lambda x: _safe_zfill(x, 8))

    result[output_col] = pv_padded + nro_padded + letras
    return result


def _render_code_builder_section(df: pd.DataFrame, prefix: str):
    """UI para construir columna de código de comprobante (diccionario AFIP integrado)."""
    p = prefix

    st.caption(
        "Formato AFIP: **punto de venta** (4 dígitos) + **número de comprobante** (8 dígitos) + **letra** del tipo. "
        "Ej: punto=1, número=34, tipo='81 - Tique Factura A' → `000100000034A`. "
        f"Diccionario integrado con {len(AFIP_TIPOS)} tipos de comprobante."
    )

    cfg = st.session_state.setdefault(f"{p}code_cfg", {})
    cols = list(df.columns)
    sc1, sc2, sc3, sc4 = st.columns(4)

    with sc1:
        cfg["tipo_col"] = st.selectbox(
            "Columna tipo (descripción)",
            cols,
            index=cols.index(cfg["tipo_col"]) if cfg.get("tipo_col") in cols else 0,
            key=f"{p}code_tipo",
            help="Ej: '81 - Tique Factura A'. Se extrae el número antes del ' - '.",
        )
    with sc2:
        cfg["pv_col"] = st.selectbox(
            "Punto de venta → 4 dígitos",
            cols,
            index=cols.index(cfg["pv_col"]) if cfg.get("pv_col") in cols else 0,
            key=f"{p}code_pv",
            help="1 → 0001 · 12 → 0012",
        )
    with sc3:
        cfg["nro_col"] = st.selectbox(
            "Número de comprobante → 8 dígitos",
            cols,
            index=cols.index(cfg["nro_col"]) if cfg.get("nro_col") in cols else 0,
            key=f"{p}code_nro",
            help="34 → 00000034 · 1000 → 00001000",
        )
    with sc4:
        cfg["output_col"] = st.text_input(
            "Nombre columna resultado",
            value=cfg.get("output_col", "Codigo_Comprobante"),
            key=f"{p}code_out",
        )

    # Preview con primera fila disponible
    if cfg.get("tipo_col") and cfg.get("pv_col") and cfg.get("nro_col") and not df.empty:
        try:
            row = df[[cfg["tipo_col"], cfg["pv_col"], cfg["nro_col"]]].dropna().iloc[0]
            tipo_s = str(row[cfg["tipo_col"]])
            m = re.match(r"^(\d+)\s*(?:-|$)", tipo_s)
            cod_raw = m.group(1) if m else "?"
            cod_norm = str(int(cod_raw)).zfill(3) if cod_raw.isdigit() else cod_raw
            letra_s = AFIP_TIPOS.get(cod_norm, "?")

            def _sz(v, w):
                try:
                    return str(int(float(str(v).replace(",", ".").strip()))).zfill(w)
                except Exception:
                    return "?" * w

            preview = f"{_sz(row[cfg['pv_col']], 4)}{_sz(row[cfg['nro_col']], 8)}{letra_s}"
            st.caption(
                f"Vista previa (primera fila): tipo=`{tipo_s}` → código={cod_norm} → letra=**{letra_s}** "
                f"| resultado: `{preview}`"
            )
        except Exception:
            pass


def _apply_code_builder(df: pd.DataFrame, col_types: dict, prefix: str) -> tuple[pd.DataFrame, dict]:
    """Aplica el código de comprobante si está configurado. Retorna (df_modificado, col_types_actualizado)."""
    cfg = st.session_state.get(f"{prefix}code_cfg", {})
    if (all(cfg.get(k) for k in ["tipo_col", "pv_col", "nro_col", "output_col"])
            and all(cfg[k] in df.columns for k in ["tipo_col", "pv_col", "nro_col"])):
        df = _build_compound_code(df, cfg["tipo_col"], cfg["pv_col"], cfg["nro_col"], cfg["output_col"])
        col_types = {**col_types, cfg["output_col"]: "str"}
    return df, col_types


def _apply_filters(df, rules, logic, col_types):
    if not rules:
        return df
    mask = None
    for r in rules:
        m = apply_rule(
            df[r["col"]], r["condition"], r.get("value"), r.get("value2"),
            dtype=col_types.get(r["col"], "str"),
        )
        if m is None:
            continue
        mask = m if mask is None else (mask & m if logic == "AND" else mask | m)

    filtered = df[mask].copy() if mask is not None else df.copy()

    # Aplicar transformaciones por regla a las filas del resultado filtrado
    for r in rules:
        transforms = r.get("transforms", [])
        if not transforms:
            continue
        rule_mask = apply_rule(
            filtered[r["col"]], r["condition"], r.get("value"), r.get("value2"),
            dtype=col_types.get(r["col"], "str"),
        )
        if rule_mask is None:
            continue
        for t in transforms:
            tcol, op, tval = t.get("col"), t.get("op", "×-1"), t.get("val")
            if not tcol or tcol not in filtered.columns:
                continue
            numeric = pd.to_numeric(filtered.loc[rule_mask, tcol], errors="coerce")
            if op == "×-1":
                filtered.loc[rule_mask, tcol] = numeric * -1
            elif op == "×" and tval is not None:
                filtered.loc[rule_mask, tcol] = numeric * tval
            elif op == "+" and tval is not None:
                filtered.loc[rule_mask, tcol] = numeric + tval
            elif op == "-" and tval is not None:
                filtered.loc[rule_mask, tcol] = numeric - tval
            elif op == "÷" and tval is not None and tval != 0:
                filtered.loc[rule_mask, tcol] = numeric / tval

    return filtered


def _apply_value_changes(df: pd.DataFrame, changes: list, col_types: dict) -> pd.DataFrame:
    """Aplica transformaciones a filas que cumplen una condición sin filtrar el resultado."""
    if not changes:
        return df
    result = df.copy()
    for ch in changes:
        cond_col = ch.get("cond_col")
        condition = ch.get("condition")
        filter_val = ch.get("filter_val", "")
        target_col = ch.get("target_col")
        op = ch.get("op", "×-1")
        op_val = ch.get("op_val")
        if not cond_col or not condition or not target_col:
            continue
        if cond_col not in result.columns or target_col not in result.columns:
            continue
        mask = apply_rule(
            result[cond_col], condition, filter_val,
            dtype=col_types.get(cond_col, "str"),
        )
        if mask is None or not mask.any():
            continue
        numeric = pd.to_numeric(result.loc[mask, target_col], errors="coerce")
        if op == "×-1":
            result.loc[mask, target_col] = numeric * -1
        elif op == "×" and op_val is not None:
            result.loc[mask, target_col] = numeric * op_val
        elif op == "+" and op_val is not None:
            result.loc[mask, target_col] = numeric + op_val
        elif op == "-" and op_val is not None:
            result.loc[mask, target_col] = numeric - op_val
        elif op == "÷" and op_val is not None and op_val != 0:
            result.loc[mask, target_col] = numeric / op_val
    return result


def _render_value_changes_section(df: pd.DataFrame, col_types: dict, prefix: str):
    """UI para definir cambios de valor condicionales (todas las filas en el resultado)."""
    key = f"{prefix}vc_defs"
    if key not in st.session_state:
        st.session_state[key] = []
    defs = st.session_state[key]

    st.caption(
        "Definí una condición y la transformación a aplicar. "
        "Las filas que la cumplan se modifican, pero el resultado incluye **todas** las filas."
    )

    VC_OPS = ["×-1", "×", "+", "-", "÷"]
    VC_OP_LABELS = {"×-1": "× -1 (cambiar signo)", "×": "×", "+": "+", "-": "-", "÷": "÷"}

    def _cond_options(dtype):
        if dtype in ("int", "float"):
            return ["equals", "not_equals", "greater", "less"]
        if dtype == "date":
            return ["equals", "before", "after"]
        return ["equals", "contains", "starts_with", "ends_with"]

    all_cols = list(df.columns)
    num_cols = [c for c in all_cols if pd.api.types.is_numeric_dtype(df[c])
                or pd.to_numeric(df[c], errors="coerce").notna().any()]

    if not num_cols:
        st.info("No hay columnas numéricas disponibles para transformar.")
        return

    to_delete = []
    for i, ch in enumerate(defs):
        c1, c2, c3, c_arr, c4, c5, c6 = st.columns([2.5, 2, 2, 0.4, 2.5, 2.5, 1])
        with c1:
            idx = all_cols.index(ch["cond_col"]) if ch.get("cond_col") in all_cols else 0
            ch["cond_col"] = st.selectbox("Columna condición", all_cols, index=idx,
                                          key=f"{prefix}vc_{i}_cc", label_visibility="collapsed")
        with c2:
            dtype = col_types.get(ch["cond_col"], "str")
            opts = _cond_options(dtype)
            labels = [RULE_LABELS.get(o, o) for o in opts]
            cur = ch.get("condition", opts[0])
            idx2 = opts.index(cur) if cur in opts else 0
            sel = st.selectbox("Condición", labels, index=idx2,
                               key=f"{prefix}vc_{i}_cond", label_visibility="collapsed")
            ch["condition"] = RULE_LABELS_INV.get(sel, sel)
        with c3:
            ch["filter_val"] = st.text_input("Valor", value=str(ch.get("filter_val", "")),
                                             key=f"{prefix}vc_{i}_fv", label_visibility="collapsed",
                                             placeholder="Valor a buscar")
        with c_arr:
            st.markdown("<div style='text-align:center;margin-top:8px;font-size:18px'>→</div>",
                        unsafe_allow_html=True)
        with c4:
            idx4 = num_cols.index(ch["target_col"]) if ch.get("target_col") in num_cols else 0
            ch["target_col"] = st.selectbox("Columna a modificar", num_cols, index=idx4,
                                            key=f"{prefix}vc_{i}_tc", label_visibility="collapsed")
        with c5:
            ch["op"] = st.selectbox("Operación", VC_OPS,
                                    index=VC_OPS.index(ch.get("op", "×-1")),
                                    format_func=lambda o: VC_OP_LABELS[o],
                                    key=f"{prefix}vc_{i}_op", label_visibility="collapsed")
            if ch["op"] != "×-1":
                ch["op_val"] = st.number_input("Valor op", value=float(ch.get("op_val") or 1.0),
                                               key=f"{prefix}vc_{i}_ov", label_visibility="collapsed")
        with c6:
            st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
            if st.button("❌", key=f"{prefix}vc_{i}_del"):
                to_delete.append(i)

    for i in reversed(to_delete):
        defs.pop(i)
        st.rerun()

    if st.button("➕ Agregar cambio", key=f"{prefix}vc_add"):
        defs.append({
            "cond_col": all_cols[0],
            "condition": _cond_options(col_types.get(all_cols[0], "str"))[0],
            "filter_val": "",
            "target_col": num_cols[0],
            "op": "×-1",
            "op_val": None,
        })
        st.rerun()


def _apply_calc_cols(df: pd.DataFrame, defs: list) -> pd.DataFrame:
    """Agrega columnas calculadas al DataFrame. Soporta N operandos encadenados."""
    if not defs:
        return df
    result = df.copy()
    for d in defs:
        name = d.get("name", "")
        c1 = d.get("col1")
        c2 = d.get("col2")
        if not name or c1 not in result.columns or c2 not in result.columns:
            continue
        try:
            op = d.get("op", "×")
            a = pd.to_numeric(result[c1], errors="coerce")
            # col2: puede ser columna o valor constante
            if d.get("col2_use_const"):
                b = float(d.get("col2_const_val") or 0)
            else:
                if c2 not in result.columns:
                    continue
                b = pd.to_numeric(result[c2], errors="coerce")
            if op == "×":
                series = a * b
            elif op == "+":
                series = a + b
            elif op == "-":
                series = a - b
            elif op == "÷":
                series = np.where(b != 0, a / b, np.nan) if not isinstance(b, float) else (a / b if b != 0 else np.nan)
            else:
                series = a * b
            # Operandos extra encadenados (columna o constante)
            for es in d.get("extra_steps", []):
                eop = es.get("op", "×")
                if es.get("use_const"):
                    e = float(es.get("const_val") or 0)
                else:
                    ec = es.get("col")
                    if not ec or ec not in result.columns:
                        continue
                    e = pd.to_numeric(result[ec], errors="coerce")
                if eop == "×":
                    series = series * e
                elif eop == "+":
                    series = series + e
                elif eop == "-":
                    series = series - e
                elif eop == "÷":
                    series = np.where(e != 0, series / e, np.nan) if not isinstance(e, float) else (series / e if e != 0 else np.nan)
            result[name] = series
        except Exception:
            logger.exception(f"Error calculando columna '{name}'")
    return result


def _numeric_cols_from(coinc: pd.DataFrame) -> list:
    """Columnas de coincidencias que son o pueden convertirse a número (int o float)."""
    cols = []
    for c in coinc.columns:
        if c == "_merge":
            continue
        if pd.api.types.is_numeric_dtype(coinc[c]):
            cols.append(c)
        else:
            # Columnas object que tienen al menos un valor numérico
            if pd.to_numeric(coinc[c], errors="coerce").notna().any():
                cols.append(c)
    return cols


def _render_calc_cols_section(coinc: pd.DataFrame):
    """UI para definir columnas calculadas sobre las coincidencias (N operandos encadenados)."""
    numeric_cols = _numeric_cols_from(coinc)

    st.caption(
        "Definí operaciones entre columnas numéricas. Podés encadenar N operandos. "
        "Ejemplo: Precio_dolar × Cantidad × IVA → Total."
    )

    if not numeric_cols:
        st.info("No hay columnas numéricas disponibles en Coincidencias.")
        return

    if "calc_definitions" not in st.session_state:
        st.session_state["calc_definitions"] = []

    defs = st.session_state["calc_definitions"]
    to_delete = []

    for i, d in enumerate(defs):
        with st.container(border=True):
            # Fila base: col1 op col2 nombre ❌
            c1, c2, c3, c4, c5 = st.columns([3, 1, 3, 3, 1])
            with c1:
                d["col1"] = st.selectbox(
                    "Col 1", numeric_cols,
                    index=numeric_cols.index(d["col1"]) if d.get("col1") in numeric_cols else 0,
                    key=f"cd_c1_{i}", label_visibility="collapsed",
                )
            with c2:
                d["op"] = st.selectbox(
                    "Op", CALC_OPS,
                    index=CALC_OPS.index(d.get("op", "×")),
                    key=f"cd_op_{i}", label_visibility="collapsed",
                )
            with c3:
                d["col2_use_const"] = st.checkbox(
                    "Valor fijo", value=d.get("col2_use_const", False),
                    key=f"cd_c2const_{i}",
                    help="Usá un número fijo como operando (ej: 100 para dividir un porcentaje).",
                )
                if d["col2_use_const"]:
                    d["col2_const_val"] = st.number_input(
                        "", value=float(d.get("col2_const_val") or 100.0),
                        key=f"cd_c2val_{i}", label_visibility="collapsed",
                    )
                else:
                    d["col2"] = st.selectbox(
                        "", numeric_cols,
                        index=numeric_cols.index(d["col2"]) if d.get("col2") in numeric_cols else 0,
                        key=f"cd_c2_{i}", label_visibility="collapsed",
                    )
            with c4:
                d["name"] = st.text_input(
                    "Nombre", value=d.get("name", f"Calculada_{i + 1}"),
                    key=f"cd_name_{i}", label_visibility="collapsed",
                    placeholder="Ej: Total",
                )
            with c5:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                if st.button("❌", key=f"cd_del_{i}"):
                    to_delete.append(i)

            # Pasos extra encadenados
            d.setdefault("extra_steps", [])
            to_del_es = []
            for ei, es in enumerate(d["extra_steps"]):
                _, ec1, ec2, _, ec5 = st.columns([3, 1, 3, 3, 1])
                with ec1:
                    es["op"] = st.selectbox(
                        "Op", CALC_OPS,
                        index=CALC_OPS.index(es.get("op", "×")),
                        key=f"cd_es_{i}_{ei}_op", label_visibility="collapsed",
                    )
                with ec2:
                    es["use_const"] = st.checkbox(
                        "Valor fijo", value=es.get("use_const", False),
                        key=f"cd_es_{i}_{ei}_const",
                        help="Usá un número fijo como operando (ej: 100 para dividir un porcentaje).",
                    )
                    if es["use_const"]:
                        es["const_val"] = st.number_input(
                            "", value=float(es.get("const_val") or 100.0),
                            key=f"cd_es_{i}_{ei}_cval", label_visibility="collapsed",
                        )
                    else:
                        es["col"] = st.selectbox(
                            "", numeric_cols,
                            index=numeric_cols.index(es["col"]) if es.get("col") in numeric_cols else 0,
                            key=f"cd_es_{i}_{ei}_col", label_visibility="collapsed",
                        )
                with ec5:
                    st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                    if st.button("❌", key=f"cd_es_{i}_{ei}_del"):
                        to_del_es.append(ei)
            for ei in reversed(to_del_es):
                d["extra_steps"].pop(ei)
                st.session_state.pop("calc_applied", None)
                st.rerun()

            # Botón agregar operando
            if st.button("➕ Agregar operando", key=f"cd_addstep_{i}"):
                d["extra_steps"].append({"op": "×", "col": numeric_cols[0]})
                st.session_state.pop("calc_applied", None)
                st.rerun()

            # Preview de la fórmula con paréntesis que reflejan el orden de evaluación
            name_p = d.get("name") or f"Calculada_{i + 1}"
            col2_str = (f"**{d.get('col2_const_val', '?')}**"
                        if d.get("col2_use_const")
                        else f"`{d.get('col2', '?')}`")
            formula = f"`{d.get('col1', '?')}` **{d.get('op', '×')}** {col2_str}"
            for es in d["extra_steps"]:
                es_str = (f"**{es.get('const_val', '?')}**"
                          if es.get("use_const")
                          else f"`{es.get('col', '?')}`")
                formula = f"({formula}) **{es.get('op', '×')}** {es_str}"
            st.caption(f"→ **{name_p}** = {formula}")

    for i in reversed(to_delete):
        defs.pop(i)
        st.session_state.pop("calc_applied", None)
        st.rerun()

    add_col, calc_col = st.columns([2, 2])
    with add_col:
        if st.button("➕ Agregar columna calculada", key="calc_add"):
            defs.append({
                "col1": numeric_cols[0],
                "op": "×",
                "col2": numeric_cols[min(1, len(numeric_cols) - 1)],
                "name": f"Calculada_{len(defs) + 1}",
                "extra_steps": [],
            })
            st.session_state.pop("calc_applied", None)
            st.rerun()
    with calc_col:
        if defs and st.button("▶️ Calcular columnas", key="btn_calc", type="primary"):
            st.session_state["calc_applied"] = True
            st.rerun()


def _reset_filters_if_table_changed(table_key: str):
    if st.session_state.get("_active_table_key") != table_key:
        st.session_state["_active_table_key"] = table_key
        for k in ("key_mappings", "compare_mappings", "concil_results",
                  "calc_definitions", "calc_applied",
                  "show_filters_m1", "show_filters_m2", "show_filters_m2b",
                  "show_vc_m1", "show_vc_m2a", "show_vc_m2b",
                  "show_code_m1", "show_code_m2a", "show_code_m2b"):
            st.session_state.pop(k, None)
        for prefix in ("m1_", "m2a_", "m2b_"):
            for k in ("rules", "current_rule", "new_rule", "logic", "vc_defs", "code_cfg"):
                st.session_state.pop(f"{prefix}{k}", None)

#Comentado por el momento ya que no esta actualizado y molesta. Luego verlo
def _help_text():
    with st.expander("ℹ️ ¿Cómo usar los filtros?"):
        st.markdown("""
**Paso a paso:**
1. Elegí una columna (Nombre, Fecha, ID…)
2. Elegí una condición (contiene, mayor que, entre…)
3. Ingresá el valor y hacé clic en **➕ Agregar regla**

**Combinar condiciones:**
- **Y** → se cumplen TODAS las condiciones
- **O** → se cumple AL MENOS una
        """)



# ── Carga de archivos ──────────────────────────────────────────────────────────
n = st.session_state.n_tables
loaded_tables = []

for row_start in range(0, n, 2):
    slots = min(2, n - row_start)
    row = st.columns(slots)
    for j in range(slots):
        idx = row_start + j
        with row[j]:
            st.markdown(f"#### Tabla {idx + 1}")
            df, _ = _render_file_loader(f"Subir Tabla {idx + 1}", f"t{idx}")
            if df is not None and not df.empty:
                loaded_tables.append({"idx": idx, "label": f"Tabla {idx + 1}", "df": df})

btn1, btn2, _ = st.columns([1, 1, 8])
with btn1:
    if n < 4 and st.button("➕ Tabla", help="Agregar una tabla más (máx. 4)"):
        st.session_state.n_tables += 1
        st.rerun()
with btn2:
    if n > 1 and st.button("➖ Tabla", help="Reducir cantidad de cargadores"):
        st.session_state.n_tables -= 1
        st.rerun()

if not loaded_tables:
    st.info("Cargá al menos un archivo para comenzar.")
    st.stop()

# ── Detección de tipos ─────────────────────────────────────────────────────────
try:
    detected_all = {
        t["idx"]: {col: detect_column_type(t["df"][col]) for col in t["df"].columns}
        for t in loaded_tables
    }
except Exception:
    logger.exception("Error al detectar tipos")
    st.error("Error al analizar columnas del archivo.")
    st.stop()

global_file_key = "_".join(
    f"{t['idx']}-{t['df'].shape[0]}x{len(t['df'].columns)}"
    for t in loaded_tables
)

with st.expander("🔧 Corregir tipos de datos detectados (opcional)"):
    st.caption("El ✏️ indica columnas modificadas manualmente.")
    type_tabs = st.tabs([t["label"] for t in loaded_tables])
    col_types_all = {}
    for tab_widget, table in zip(type_tabs, loaded_tables):
        with tab_widget:
            col_types_all[table["idx"]] = _render_type_overrides(
                table["df"], detected_all[table["idx"]],
                f"overrides_t{table['idx']}", global_file_key,
            )

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# MODO 1: Una sola tabla — filtrado
# ══════════════════════════════════════════════════════════════════════════════
if len(loaded_tables) == 1:
    table = loaded_tables[0]
    df = table["df"]
    col_types = col_types_all[table["idx"]]

    _reset_filters_if_table_changed(f"single-t{table['idx']}-{global_file_key}")

    st.subheader("🔍 Modo filtro — una sola planilla")
    st.info(
        f"📋 **{len(df):,} registros · {len(df.columns)} columnas** cargados. "
        "Podés exportar todos los datos o aplicar filtros para quedarte con un subconjunto."
    )

    # ── Código de comprobante ──────────────────────────────────────────────────
    show_code = st.checkbox(
        "🔤 Construir columna de código de comprobante (opcional)",
        key="show_code_m1",
        help="Genera un código combinando punto de venta, número y letra del tipo (formato AFIP).",
    )
    if show_code:
        with st.container(border=True):
            st.markdown("#### 🔤 Construir código de comprobante")
            _render_code_builder_section(df, "m1_")
    if show_code:
        df, col_types = _apply_code_builder(df, col_types, "m1_")

    # ── Cambios de valor condicionales ────────────────────────────────────────
    show_vc = st.checkbox(
        "🔄 Cambios de valor condicionales (opcional)",
        key="show_vc_m1",
        help="Modificá valores en columnas según una condición. Todas las filas aparecen en el resultado.",
    )
    if show_vc:
        with st.container(border=True):
            st.markdown("#### 🔄 Cambios de valor por condición")
            _render_value_changes_section(df, col_types, "m1_")

    show_filters = st.checkbox(
        "🔍 Filtrar registros antes de exportar (opcional)",
        key="show_filters_m1",
        help="Activá para incluir solo ciertos registros en el resultado. Sin filtros se exporta todo.",
    )

    rules, logic = [], "AND"
    if show_filters:
        with st.container(border=True):
            rules, logic = rule_builder(df, col_types, state_prefix="m1_")
        if rules:
            st.success(f"🔍 {len(rules)} filtro(s) activo(s) con lógica '{logic}'")
        else:
            st.caption("Agregá al menos una regla arriba para filtrar.")

    btn_label = "Aplicar y ver resultado" if (show_vc or show_filters) else "📊 Ver todos los registros"
    if st.button(btn_label, type="primary"):
        df_work = _apply_value_changes(df, st.session_state.get("m1_vc_defs", []), col_types)
        resultado = _apply_filters(df_work, rules, logic, col_types) if (show_filters and rules) else df_work
        n_changed = len(st.session_state.get("m1_vc_defs", []))
        if show_filters and rules:
            st.success(f"Mostrando {len(resultado):,} de {len(df):,} registros" +
                       (f" · {n_changed} cambio(s) aplicado(s)" if n_changed else ""))
        else:
            st.success(f"Mostrando todos los {len(resultado):,} registros" +
                       (f" · {n_changed} cambio(s) de valor aplicado(s)" if n_changed else ""))
        st.dataframe(resultado, use_container_width=True)
        try:
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                resultado.to_excel(writer, index=False, sheet_name="Resultados")
            st.download_button(
                "📥 Descargar resultado", output.getvalue(), "resultado_filtrado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception:
            logger.exception("Error al generar descarga")
            st.error("Error al generar el archivo de descarga.")

# ══════════════════════════════════════════════════════════════════════════════
# MODO 2: Dos o más tablas — conciliación
# ══════════════════════════════════════════════════════════════════════════════
else:
    labels = [t["label"] for t in loaded_tables]
    table_by_label = {t["label"]: t for t in loaded_tables}

    # ── Paso 1: Seleccionar tablas ─────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("#### Seleccionar tablas a conciliar")
        c1, c2 = st.columns(2)
        with c1:
            sel_a = st.selectbox("Tabla A — origen / referencia", labels, key="sel_a")
        with c2:
            opts_b = [l for l in labels if l != sel_a]
            cur_b = st.session_state.get("sel_b_val")
            default_b = cur_b if cur_b in opts_b else opts_b[0]
            sel_b = st.selectbox(
                "Tabla B — destino / comparación", opts_b,
                index=opts_b.index(default_b), key="sel_b_val",
            )

    table_a = table_by_label[sel_a]
    table_b = table_by_label[sel_b]
    df_a = table_a["df"]
    df_b = table_b["df"]
    col_types_a = col_types_all[table_a["idx"]]
    col_types_b = col_types_all[table_b["idx"]]
    col_types = col_types_a

    active_key = f"recon-{table_a['idx']}-{table_b['idx']}-{global_file_key}"
    _reset_filters_if_table_changed(active_key)

    if df_a.empty or df_b.empty:
        st.error("Una de las tablas seleccionadas está vacía.")
        st.stop()

    # ── Código de comprobante (opcional, antes del mapeo) ──────────────────────
    show_code_a = st.checkbox(
        f"🔤 Construir código de comprobante en {sel_a} (opcional)",
        key="show_code_m2a",
        help="Genera un código combinando punto de venta, número y letra del tipo (formato AFIP).",
    )
    if show_code_a:
        with st.container(border=True):
            st.markdown(f"#### 🔤 Construir código — {sel_a}")
            _render_code_builder_section(df_a, "m2a_")
        df_a, col_types_a = _apply_code_builder(df_a, col_types_a, "m2a_")
        col_types = col_types_a

    show_code_b = st.checkbox(
        f"🔤 Construir código de comprobante en {sel_b} (opcional)",
        key="show_code_m2b",
        help="Genera un código combinando punto de venta, número y letra del tipo (formato AFIP).",
    )
    if show_code_b:
        with st.container(border=True):
            st.markdown(f"#### 🔤 Construir código — {sel_b}")
            _render_code_builder_section(df_b, "m2b_")
        df_b, col_types_b = _apply_code_builder(df_b, col_types_b, "m2b_")

    # ── Paso 2: Mapeo de columnas (requerido) ──────────────────────────────────
    mapper_key = f"a{table_a['idx']}-b{table_b['idx']}-{df_a.shape}-{df_b.shape}-{list(df_a.columns)[:5]}"
    key_mappings, compare_mappings = column_mapper(df_a, df_b, mapper_key)

    # ── Cambios de valor condicionales ────────────────────────────────────────
    show_vc_a = st.checkbox(
        f"🔄 Cambios de valor en {sel_a} antes de conciliar (opcional)",
        key="show_vc_m2a",
        help="Modificá valores en columnas de Tabla A según una condición. Todas las filas entran a la conciliación.",
    )
    if show_vc_a:
        with st.container(border=True):
            st.markdown(f"#### 🔄 Cambios de valor en {sel_a}")
            _render_value_changes_section(df_a, col_types_a, "m2a_")

    show_vc_b = st.checkbox(
        f"🔄 Cambios de valor en {sel_b} antes de conciliar (opcional)",
        key="show_vc_m2b",
        help="Modificá valores en columnas de Tabla B según una condición. Todas las filas entran a la conciliación.",
    )
    if show_vc_b:
        with st.container(border=True):
            st.markdown(f"#### 🔄 Cambios de valor en {sel_b}")
            _render_value_changes_section(df_b, col_types_b, "m2b_")

    # ── Paso 3: Filtros opcionales ─────────────────────────────────────────────
    show_filters_a = st.checkbox(
        f"🔍 Filtrar registros de {sel_a} antes de conciliar (opcional)",
        key="show_filters_m2",
        help=f"Activá para incluir solo ciertos registros de {sel_a} en la conciliación.",
    )
    rules_a, logic_a = [], "AND"
    if show_filters_a:
        with st.container(border=True):
            rules_a, logic_a = rule_builder(df_a, col_types_a, state_prefix="m2a_")
        if rules_a:
            st.success(f"🔍 {len(rules_a)} filtro(s) en {sel_a} con lógica '{logic_a}'")
        else:
            st.caption(f"Agregá al menos una regla arriba para filtrar registros de {sel_a}.")

    show_filters_b = st.checkbox(
        f"🔍 Filtrar registros de {sel_b} antes de conciliar (opcional)",
        key="show_filters_m2b",
        help=f"Activá para incluir solo ciertos registros de {sel_b} en la conciliación.",
    )
    rules_b, logic_b = [], "AND"
    if show_filters_b:
        with st.container(border=True):
            rules_b, logic_b = rule_builder(df_b, col_types_b, state_prefix="m2b_")
        if rules_b:
            st.success(f"🔍 {len(rules_b)} filtro(s) en {sel_b} con lógica '{logic_b}'")
        else:
            st.caption(f"Agregá al menos una regla arriba para filtrar registros de {sel_b}.")

    # ── Ejecutar ───────────────────────────────────────────────────────────────
    if st.button("▶️ Ejecutar Conciliación", type="primary") and key_mappings:
        # Aplicar cambios de valor antes de filtrar
        df_a_work = _apply_value_changes(df_a, st.session_state.get("m2a_vc_defs", []), col_types_a)
        df_b_work = _apply_value_changes(df_b, st.session_state.get("m2b_vc_defs", []), col_types_b)
        df_a_filtered = _apply_filters(df_a_work, rules_a, logic_a, col_types_a)
        df_b_filtered = _apply_filters(df_b_work, rules_b, logic_b, col_types_b)

        bad_a = [m["col_a"] for m in key_mappings + compare_mappings if m["col_a"] not in df_a_filtered.columns]
        bad_b = [m["col_b"] for m in key_mappings + compare_mappings if m["col_b"] not in df_b_filtered.columns]
        if bad_a:
            st.error(f"Columnas no encontradas en {sel_a}: {', '.join(bad_a)}")
            st.stop()
        if bad_b:
            st.error(f"Columnas no encontradas en {sel_b}: {', '.join(bad_b)}")
            st.stop()

        try:
            coincidencias, solo_a, solo_b, diferencias = conciliar(
                df_a_filtered, df_b_filtered, key_mappings, compare_mappings,
                keep_b_keys=st.session_state.get("keep_b_keys", False),
            )
        except Exception:
            logger.exception("Error durante la conciliación")
            st.error("Error durante la conciliación.")
            st.stop()

        st.session_state["concil_results"] = {
            "coincidencias": coincidencias,
            "solo_a": solo_a,
            "solo_b": solo_b,
            "diferencias": diferencias,
            "df_a_filtered": df_a_filtered,
            "df_b_filtered": df_b_filtered,
        }
        st.session_state.pop("calc_definitions", None)
        st.session_state.pop("calc_applied", None)

    # ── Resultados (persisten entre reruns) ────────────────────────────────────
    if "concil_results" in st.session_state:
        res = st.session_state["concil_results"]
        coincidencias = res["coincidencias"]
        solo_a       = res["solo_a"]
        solo_b       = res["solo_b"]
        diferencias  = res["diferencias"]
        df_a_filtered = res["df_a_filtered"]
        df_b_filtered = res.get("df_b_filtered", df_b)

        # ── Métricas ──────────────────────────────────────────────────────────
        st.success("✅ Conciliación completada.")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("✅ Coincidencias", len(coincidencias))
        m2.metric(f"📌 Solo en {sel_a}", len(solo_a))
        m3.metric(f"📌 Solo en {sel_b}", len(solo_b))
        m4.metric("⚠️ Diferencias", len(diferencias))

        # ── Columnas calculadas (opcional) ────────────────────────────────────
        has_defs = bool(st.session_state.get("calc_definitions", []))
        with st.expander("🔢 Agregar columnas calculadas al resultado (opcional)", expanded=has_defs):
            _render_calc_cols_section(coincidencias)

        calc_defs = st.session_state.get("calc_definitions", [])
        if calc_defs and st.session_state.get("calc_applied"):
            coinc_con_calcs = _apply_calc_cols(coincidencias, calc_defs)
        else:
            coinc_con_calcs = coincidencias

        # ── Vista previa ───────────────────────────────────────────────────────
        key_col_names = [m["col_a"] for m in key_mappings]

        with st.expander("📊 Vista previa de resultados", expanded=True):
            result_tabs = st.tabs([
                "✅ Coincidencias",
                f"📌 Solo en {sel_a}",
                f"📌 Solo en {sel_b}",
                "⚠️ Diferencias",
            ])
            coinc_preview = coinc_con_calcs.drop(columns=["_merge"], errors="ignore")
            with result_tabs[0]:
                st.dataframe(coinc_preview.head(10), use_container_width=True)
            with result_tabs[1]:
                if not solo_a.empty:
                    st.dataframe(solo_a.head(10), use_container_width=True)
                else:
                    st.info(f"No hay registros exclusivos de {sel_a}.")
            with result_tabs[2]:
                if not solo_b.empty:
                    st.dataframe(solo_b.head(10), use_container_width=True)
                else:
                    st.info(f"No hay registros exclusivos de {sel_b}.")
            with result_tabs[3]:
                if not diferencias.empty:
                    st.dataframe(diferencias.head(10), use_container_width=True)
                else:
                    st.info("No se encontraron diferencias en las columnas seleccionadas.")

        # ── Configuración de descarga ──────────────────────────────────────────
        with st.expander("⚙️ Configuración de descarga (opcional)", expanded=False):
            dl_filename = st.text_input(
                "Nombre del archivo (sin extensión)",
                value="conciliacion_resultado",
                key="dl_filename",
            )

            st.markdown("**Hojas a incluir** (Coincidencias siempre se incluye):")
            hc1, hc2, hc3, hc4, hc5 = st.columns(5)
            inc_diffs  = hc1.checkbox("Diferencias",          value=True,  key="dl_diffs")
            show_solo_a = hc2.checkbox(f"Solo {sel_a[:16]}",  value=False, key="dl_solo_a",
                                       help=f"Filas de {sel_a} sin coincidencia en {sel_b}")
            show_solo_b = hc3.checkbox(f"Solo {sel_b[:16]}",  value=False, key="dl_solo_b",
                                       help=f"Filas de {sel_b} sin coincidencia en {sel_a}")
            inc_orig_a  = hc4.checkbox(f"{sel_a[:14]} orig.", value=False, key="dl_orig_a",
                                       help=f"Incluir tabla {sel_a} completa tal como fue cargada")
            inc_orig_b  = hc5.checkbox(f"{sel_b[:14]} orig.", value=False, key="dl_orig_b",
                                       help=f"Incluir tabla {sel_b} completa tal como fue cargada")

            all_coinc_cols = [c for c in coinc_con_calcs.columns if c != "_merge"]
            optional_cols  = [c for c in all_coinc_cols if c not in key_col_names]

            if key_col_names:
                fixed_label = ", ".join(f"`{c}`" for c in key_col_names if c in all_coinc_cols)
                st.markdown(f"**Columnas de coincidencia** (siempre incluidas): {fixed_label}")

            if optional_cols:
                dl_extra_cols = st.multiselect(
                    "Columnas adicionales a incluir en Coincidencias y Diferencias:",
                    optional_cols,
                    default=optional_cols,
                    key="dl_extra_cols",
                )
            else:
                dl_extra_cols = []

        # Columnas finales para la hoja Coincidencias
        fixed_cols = [c for c in key_col_names if c in coinc_con_calcs.columns]
        final_coinc_cols = fixed_cols + [c for c in dl_extra_cols if c in coinc_con_calcs.columns]
        coinc_export = coinc_con_calcs.drop(columns=["_merge"], errors="ignore")
        if final_coinc_cols:
            coinc_export = coinc_export[[c for c in final_coinc_cols if c in coinc_export.columns]]

        # ── Descarga ───────────────────────────────────────────────────────────
        try:
            output = BytesIO()
            safe_name = (dl_filename.strip() or "conciliacion_resultado") + ".xlsx"
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                coinc_export.to_excel(writer, index=False, sheet_name="Coincidencias")
                if inc_diffs and not diferencias.empty:
                    selected_set = set(dl_extra_cols)
                    dif_cols = [
                        c for c in diferencias.columns
                        if c in fixed_cols
                        or c == "Columnas_con_diferencias"
                        or c in selected_set
                    ]
                    dif_export = diferencias[dif_cols] if dif_cols else diferencias
                    dif_export.to_excel(writer, index=False, sheet_name="Diferencias")
                    _highlight_diff_cells(writer.sheets["Diferencias"])
                if show_solo_a:
                    solo_a.drop(columns=["_merge"], errors="ignore").to_excel(
                        writer, index=False, sheet_name=f"Solo {sel_a}"[:31]
                    )
                if show_solo_b:
                    solo_b.drop(columns=["_merge"], errors="ignore").to_excel(
                        writer, index=False, sheet_name=f"Solo {sel_b}"[:31]
                    )
                if inc_orig_a:
                    df_a_filtered.to_excel(writer, index=False, sheet_name=f"{sel_a} Original"[:31])
                if inc_orig_b:
                    df_b_filtered.to_excel(writer, index=False, sheet_name=f"{sel_b} Original"[:31])

            st.download_button(
                "📥 Descargar Reporte de Conciliación",
                output.getvalue(),
                safe_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception:
            logger.exception("Error al generar descarga")
            st.error("Error al generar el archivo de descarga.")
